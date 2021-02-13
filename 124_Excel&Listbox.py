"""
Tkinter - Codemy.com #124 : How To Add Excel Spreadsheet Columns To Listbox - Python Tkinter GUI Tutorial #124
Lien : https://www.youtube.com/watch?v=Kcduy4fPZWg&list=PLCC34OHNcOtoC6GglhF3ncJ5rLwQrLGnV&index=124

Dans ce programme on apprend à insérer dans le widget listbox des données issues d'Excel

Pour se mettre en lien avec Excel, il faut le module openpyxl qui n'est pas automatiquement installé en tant que module
comme par exemples math et random

Éditeur : Laurent REYNAUD
Date : 23-12-20
"""

from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.iconbitmap('images/Logo.ico')
root.title('Titre !')
root.geometry('400x300')


def select(e):
    """Fonction permettant d'afficher la ligne sélectionnée dans la listbox"""
    my_label.config(text=my_listbox.get(ANCHOR))  # ANCHOR permet d'afficher la valeur attribuée à la ligne sélectionnée


"""Assignation d'une variable au classeur Excel"""
wb = Workbook()

"""Chargement du classeur Excel"""
wb = load_workbook('pieces/name_color.xlsx')

"""Assignation d'une variable à la feuille active d'Excel"""
ws = wb.active

"""Assignation d'une variable pour extraire les données des colonnes A et B de la feuille active d'Excel"""
col_a = ws['A']
col_b = ws['B']

"""ListBox"""
my_listbox = Listbox(root, width=45)
my_listbox.pack(pady=20)

"""Insertion des données d'Excel dans la listbox"""
for item in col_a:
    my_listbox.insert(END, item.value)

"""Etiquette"""
my_label = Label(root, text='Sélectionner une donnée...', font='Helvetica 18')
my_label.pack(pady=20)

"""Binding pour générer un évènement dès qu'on clique sur une donnée"""
my_listbox.bind('<ButtonRelease-1>', select)

root.mainloop()
