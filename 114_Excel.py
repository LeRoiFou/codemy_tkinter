"""
Tkinter - Codemy.com #114 : Using Excel Spreadsheets With Tkinter and Openpyxl
Lien : https://www.youtube.com/watch?v=l6-HG0FJPsQ&list=PLCC34OHNcOtoC6GglhF3ncJ5rLwQrLGnV&index=114

Dans ce programme on apprend à lier les données d'Excel avec python, en récupérant les données d'Excel, mais également
en ajoutant des données de Python vers Excel

Package installé : openpyxl
Pour travailler avec un fichier Excel, enregistrer le fichier CALC sous le format xlsx (Excel 2007-365)

Les données provenant d'Excel sont des tuples : il faut donc recourir à une boucle pour récupérer ces données

Éditeur : Laurent REYNAUD
Date : 20-12-20
"""

from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.iconbitmap('images/Logo.ico')
root.title('Titre !')
root.geometry('500x800')


def get_a():
    """Fonction permettant de récupérer les données de la colonne A d'Excel"""
    my_list = ''  # présentation des données sous la forme d'une str
    for cell in column_a:
        my_list = f"{my_list + str(cell.value)}\n"
    label_a.config(text=my_list)


def get_b():
    """Fonction permettant de récupérer les données de la colonne B d'Excel"""
    my_list = ''  # présentation des données sous la forme d'une str
    for cell in column_b:
        my_list = f"{my_list + str(cell.value)}\n"
    label_b.config(text=my_list)


"""Assignation d'une variable du classeur"""
wb = Workbook()

"""Chargement du classeur existant"""
wb = load_workbook('pieces/Pizzas.xlsx')

"""Assignation d'une variable de la feuille active"""
ws = wb.active

"""Assignation d'une variable pour extraire les données des colonnes A et B"""
column_a = ws['A']
column_b = ws['B']

"""Configuration du bouton d'exécution permettant de récupérer les données de la colonne A"""
ba = Button(root, text='Obtenir les données de la colonne A', command=get_a)
ba.pack(pady=20)

"""Affichage des données de la colonne A"""
label_a = Label(root, text='')
label_a.pack(pady=20)

"""Configuration du bouton d'exécution permettant de récupérer les données de la colonne B"""
bb = Button(root, text='Obtenir les données de la colonne B', command=get_b)
bb.pack(pady=20)

"""Affichage des données de la colonne B"""
label_b = Label(root, text='')
label_b.pack(pady=20)

"""Ajout de données supplémentaires de Python à Excel"""
ws['A8'] = 'Fred'  # ajout de la donnée dans la cellule A8
ws['B8'] = 'Nutella'  # ajout de la donnée dans la cellule B8

"""Sauvegarde des données ajoutées dans un nouveau classeur Excel"""
wb.save('pieces/Pizzas2.xlsx')

root.mainloop()
