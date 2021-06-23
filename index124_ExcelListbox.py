"""
Tkinter - Codemy.com #124 : 
How To Add Excel Spreadsheet Columns To Listbox 
- Python Tkinter GUI Tutorial #124
Lien : https://www.youtube.com/watch?v=Kcduy4fPZWg&list=PLCC34OHNcOtoC6GglhF3ncJ5rLwQrLGnV&index=124

Dans ce programme on apprend à insérer dans 
le widget listbox des données issues d'Excel

Pour se mettre en lien avec Excel, il faut le module openpyxl 
qui n'est pas automatiquement installé en tant que module
comme par exemples math et random

Éditeur : Laurent REYNAUD
Date : 23-12-20
"""

from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

class Gui:
    
    def __init__(self, root):
        "Constructeur de la classe"
        
        # Configuration de la fenêtre principale
        self.root = root
        root.iconbitmap('images/Logo.ico')
        root.title('Titre !')
        root.geometry('400x300')
        
        # Appel des méthodes ci-après
        self.my_excel()
        self.widgets()
        
    def my_excel(self):
        "Accès aux données d'Excel"
        
        """Assignation d'une variable au classeur Excel"""
        wb = Workbook()

        """Chargement du classeur Excel"""
        wb = load_workbook('pieces/name_color.xlsx')

        """Assignation d'une variable à la feuille active d'Excel"""
        ws = wb.active

        """Assignation d'une variable pour extraire les données 
        des colonnes A et B de la feuille active d'Excel"""
        self.col_a = ws['A']
        self.col_b = ws['B']
        
    def widgets(self):
        "Configuration des widgets"
        
        """ListBox"""
        self.my_listbox = Listbox(root, width=45)
        self.my_listbox.pack(pady=20)

        """Insertion des données d'Excel dans la listbox"""
        for item in self.col_a:
            self.my_listbox.insert(END, item.value)

        """Etiquette"""
        self.my_label = Label(
            root, 
            text='Sélectionner une donnée...', 
            font='Helvetica 18')
        self.my_label.pack(pady=20)

        """Binding pour générer un évènement 
        dès qu'on clique sur une donnée"""
        self.my_listbox.bind('<ButtonRelease-1>', self.select)
        
    def select(self, e):
        """Méthode permettant d'afficher 
        la ligne sélectionnée dans la listbox"""
        
        # ANCHOR permet d'afficher la valeur attribuée à la ligne sélectionnée
        self.my_label.config(text=self.my_listbox.get(ANCHOR))  


if __name__ == '__main__':
    root = Tk()
    gui = Gui(root)
    root.mainloop()
